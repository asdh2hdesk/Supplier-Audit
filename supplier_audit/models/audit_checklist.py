from odoo import models, fields, api, _
from odoo.exceptions import UserError
from openpyxl import load_workbook
from io import BytesIO
import base64
import os

# ALLOWED_CATEGORIES = [
#     'Management',
#     'Manufacturing',
#     'Production Readiness',
#     'Quality Assurance & Process'
# ]

class AuditChecklist(models.Model):
    _name = 'audit.checklist'
    _description = 'Audit Checklist Template'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char('Checklist Name', required=True, tracking=True)
    code = fields.Char('Checklist Code', tracking=True)
    description = fields.Text('Description')
    version = fields.Char('Version', default='1.0', tracking=True)
    active = fields.Boolean('Active', default=True, tracking=True)

    question_ids = fields.One2many('audit.checklist.question', 'checklist_id', string='Questions')
    total_questions = fields.Integer('Total Questions', compute='_compute_question_count', store=True)
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)

    uploaded_excel_file = fields.Binary(string="Upload Excel File", required=True)
    uploaded_file_name = fields.Char(string="File Name")

    @api.depends('question_ids')
    def _compute_question_count(self):
        for record in self:
            record.total_questions = len(record.question_ids)

    def copy(self, default=None):
        default = dict(default or {})
        default.update({
            'name': _('%s (Copy)') % self.name,
            'version': self.version + ' (Copy)',
        })
        return super().copy(default)

    # def _get_or_create_allowed_categories(self):
    #     category_map = {}
    #     for cat_name in ALLOWED_CATEGORIES:
    #         category = self.env['audit.question.category'].search([('name', '=', cat_name)], limit=1)
    #         if not category:
    #             category = self.env['audit.question.category'].create({'name': cat_name})
    #         category_map[cat_name] = category
    #     return category_map

    def action_upload_questions(self):
        self.ensure_one()

        if not self.uploaded_excel_file:
            raise UserError(_("Please upload an Excel file."))

        try:
            decoded_file = base64.b64decode(self.uploaded_excel_file)
            workbook = load_workbook(filename=BytesIO(decoded_file), data_only=True)
        except Exception as e:
            raise UserError(f"Error reading Excel file: {str(e)}")

        # Normalization map (optional use if category is found)
        CATEGORY_NORMALIZATION = {
            'management': 'Management',
            'manufacturing': 'Manufacturing',
            'production readiness': 'Production Readiness',
            'production': 'Production Readiness',
            'quality': 'Quality Assurance & Process',
            'quality assurance': 'Quality Assurance & Process',
            'qa': 'Quality Assurance & Process',
        }

        # Build existing category map (only used if category is found)
        category_map = {}
        categories = self.env['audit.question.category'].search([])
        for cat in categories:
            category_map[cat.name.lower()] = cat.id

        questions = []

        for sheet in workbook.worksheets:
            headers = {}
            header_row_index = None

            # Auto-detect header row
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                lower_row = [str(cell).strip().lower() if cell else '' for cell in row]
                if 'question' in lower_row:
                    headers = {str(cell).strip().lower(): i for i, cell in enumerate(row) if cell}
                    header_row_index = row_idx
                    break

            if not headers or header_row_index is None:
                raise UserError(_("Could not detect a valid header row. Please include a 'Question' column."))

            for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
                if not row or not row[headers.get('question', -1)]:
                    continue

                def get_cell(col_name):
                    idx = headers.get(col_name.lower())
                    return str(row[idx]).strip() if idx is not None and row[idx] else ''

                # Get category from 'category' or 'category_id' if present
                category_col = 'category' if 'category' in headers else 'category_id' if 'category_id' in headers else ''
                category_raw = get_cell(category_col).lower() if category_col else ''
                normalized_name = CATEGORY_NORMALIZATION.get(category_raw, category_raw)
                category_id = category_map.get(normalized_name.lower()) if normalized_name else False

                status = get_cell('status')
                if status not in ['0', '1', '2', '3']:
                    status = '3'

                question_vals = {
                    'name': get_cell('question'),
                    'evidence': get_cell('evidence required'),
                    'scoring_criteria': get_cell('scoring criteria'),
                    'status': status,
                    'observation': get_cell('observation'),
                    'action': get_cell('action'),
                }

                if category_id:
                    question_vals['category_id'] = category_id

                questions.append((0, 0, question_vals))

        if not questions:
            raise UserError(_("No valid questions found in the uploaded file."))

        self.question_ids = questions


class AuditChecklistQuestion(models.Model):
    _name = 'audit.checklist.question'
    _description = 'Audit Checklist Question'
    _order = 'sequence, id'

    sequence = fields.Integer('Sequence Number', default=10)
    sl_no = fields.Integer('Sl.No', compute='_compute_serial_no', store=True)
    name = fields.Text('Question', required=True)
    evidence = fields.Text('Evidence')
    scoring_criteria = fields.Text('Scoring Criteria')
    status = fields.Selection([
        ('0', '0'), ('1', '1'), ('2', '2'), ('3', '3')
    ], string='Score', default='3')
    observation = fields.Text('Observation')
    action = fields.Text('Action')

    checklist_id = fields.Many2one('audit.checklist', string='Checklist', ondelete='cascade')
    category_id = fields.Many2one(
        'audit.question.category',
        string='Category',
        domain="[('name', 'in', ['Management', 'Manufacturing', 'Production Readiness', 'Quality Assurance & Process'])]"
    )

    @api.depends('sequence', 'checklist_id')
    def _compute_serial_no(self):
        for checklist in self.mapped('checklist_id'):
            for i, line in enumerate(checklist.question_ids, start=1):
                line.sl_no = i

class AuditQuestionCategory(models.Model):
    _name = 'audit.question.category'
    _description = 'Audit Question Category'
    _order = 'sequence, name'

    name = fields.Char('Category Name', required=True)
    description = fields.Text('Description')
    sequence = fields.Integer('Sequence', default=10)
    color = fields.Integer('Color Index')

    question_ids = fields.One2many('audit.checklist.question', 'category_id', string='Questions')
    question_count = fields.Integer('Question Count', compute='_compute_question_count', store=True)

    @api.depends('question_ids')
    def _compute_question_count(self):
        for record in self:
            record.question_count = len(record.question_ids)

    @api.model
    def delete_junk_categories(self):
        allowed = ALLOWED_CATEGORIES
        junk_categories = self.search([
            ('name', '!=', False),
            ('name', 'not in', allowed)
        ])
        # Only delete categories that are purely numbers (e.g. "10", "11")
        junk_categories = junk_categories.filtered(lambda c: c.name.strip().isdigit())
        junk_categories.unlink()